from __future__ import annotations

from pathlib import Path
from typing import Iterable

import openpyxl
from PIL import Image, ImageChops, ImageDraw, ImageFont


ROOT = Path(__file__).resolve().parent
OUT_DIR = ROOT.parents[1] / "outputs" / "sponsor-list"

CANVAS_W = 1584
CANVAS_H = 2016

NAVY = (42, 54, 90)
GOLD = (226, 197, 111)
BRONZE = (183, 132, 82)
BLUE_BLOCK = (87, 111, 172)
GREY = (145, 145, 145)
WHITE = (255, 255, 255)


def font(path: str, size: int) -> ImageFont.FreeTypeFont:
    try:
        return ImageFont.truetype(path, size)
    except OSError:
        return ImageFont.truetype("/Library/Fonts/Arial Unicode.ttf", size)


FONT_REG = "/System/Library/Fonts/Avenir Next.ttc"
FONT_COND = "/System/Library/Fonts/Avenir Next Condensed.ttc"
FONT_HELV = "/System/Library/Fonts/HelveticaNeue.ttc"


def text_size(draw: ImageDraw.ImageDraw, text: str, fnt: ImageFont.ImageFont) -> tuple[int, int]:
    box = draw.textbbox((0, 0), text, font=fnt)
    return box[2] - box[0], box[3] - box[1]


def draw_centered(
    draw: ImageDraw.ImageDraw,
    xy: tuple[int, int],
    text: str,
    fnt: ImageFont.ImageFont,
    fill: tuple[int, int, int],
) -> None:
    x, y = xy
    w, h = text_size(draw, text, fnt)
    draw.text((x - w / 2, y - h / 2), text, font=fnt, fill=fill)


def tracking_width(draw: ImageDraw.ImageDraw, text: str, fnt: ImageFont.ImageFont, spacing: int) -> int:
    if not text:
        return 0
    return sum(text_size(draw, ch, fnt)[0] for ch in text) + spacing * (len(text) - 1)


def draw_tracking_centered(
    draw: ImageDraw.ImageDraw,
    center_x: int,
    y: int,
    text: str,
    fnt: ImageFont.ImageFont,
    fill: tuple[int, int, int],
    spacing: int,
) -> None:
    width = tracking_width(draw, text, fnt, spacing)
    x = center_x - width / 2
    for ch in text:
        draw.text((x, y), ch, font=fnt, fill=fill)
        x += text_size(draw, ch, fnt)[0] + spacing


def trim_image(img: Image.Image, threshold: int = 248) -> Image.Image:
    rgba = img.convert("RGBA")
    alpha = rgba.getchannel("A")
    if alpha.getbbox():
        box = alpha.getbbox()
    else:
        bg = Image.new("RGB", rgba.size, WHITE)
        rgb = rgba.convert("RGB")
        diff = ImageChops.difference(rgb, bg).convert("L")
        mask = diff.point(lambda p: 255 if p < threshold else 0)
        box = ImageChops.invert(mask).getbbox()
    if not box:
        return rgba
    cropped = rgba.crop(box)
    return cropped


def fit_image(path: Path, max_w: int, max_h: int) -> Image.Image:
    img = trim_image(Image.open(path))
    img.thumbnail((max_w, max_h), Image.Resampling.LANCZOS)
    return img


def paste_centered(base: Image.Image, img: Image.Image, box: tuple[int, int, int, int]) -> None:
    x1, y1, x2, y2 = box
    x = x1 + (x2 - x1 - img.width) // 2
    y = y1 + (y2 - y1 - img.height) // 2
    if img.mode == "RGBA":
        base.alpha_composite(img, (x, y))
    else:
        base.paste(img.convert("RGBA"), (x, y))


def read_sponsors() -> tuple[list[dict[str, str]], list[str]]:
    wb = openpyxl.load_workbook(ROOT / "Sponsors_from_drive.xlsx", data_only=True)
    ws = wb[wb.sheetnames[0]]
    gold: list[dict[str, str]] = []
    bronze: list[str] = []
    section = ""
    for row in ws.iter_rows(values_only=True):
        first = row[0]
        if first == "Gold Sponsors":
            section = "gold"
            continue
        if first == "Bronze Sponsor":
            section = "bronze"
            continue
        if first == "Sponsor Name" or first is None:
            continue
        if section == "gold":
            batch = str(int(row[4])) if isinstance(row[4], (int, float)) else str(row[4] or "")
            gold.append(
                {
                    "name": str(row[0]),
                    "alumni": str(row[3] or ""),
                    "batch": batch,
                }
            )
        elif section == "bronze":
            bronze.append(str(first))
    return gold, bronze


def batch_short(batch: str) -> str:
    digits = "".join(ch for ch in batch if ch.isdigit())
    return digits[-2:] if digits else batch


def label_for(sponsor: dict[str, str]) -> str:
    alumni = sponsor["alumni"].upper()
    return f"{alumni} '{batch_short(sponsor['batch'])}" if sponsor["batch"] else alumni


def logo_for(name: str) -> Path:
    key = name.lower()
    if "alpha" in key:
        return ROOT / "alpha_omega_logo.png"
    if "aimdrive" in key:
        return ROOT / "aimdrive_logo.png"
    if "meta mind" in key:
        return ROOT / "mmgc_logo.png"
    if "learn for life" in key:
        return ROOT / "learn_for_life_logo.jpg"
    raise ValueError(f"No logo mapping for {name}")


def draw_vertical_band(
    base: Image.Image,
    label: str,
    x: int,
    y1: int,
    y2: int,
    color: tuple[int, int, int],
    font_size: int,
) -> None:
    draw = ImageDraw.Draw(base)
    draw.rectangle((x, y1, x + 58, y2), fill=color)
    fnt = font(FONT_COND, font_size)
    temp = Image.new("RGBA", (y2 - y1, 80), (255, 255, 255, 0))
    td = ImageDraw.Draw(temp)
    draw_tracking_centered(td, temp.width // 2, 5, label, fnt, WHITE, 1)
    temp = temp.rotate(270, expand=True)
    base.alpha_composite(temp, (x + (58 - temp.width) // 2, y1 + (y2 - y1 - temp.height) // 2))


def draw_sponsor_cell(
    base: Image.Image,
    sponsor: dict[str, str],
    box: tuple[int, int, int, int],
    max_logo_h: int,
) -> None:
    draw = ImageDraw.Draw(base)
    x1, y1, x2, y2 = box
    logo_box = (x1 + 24, y1, x2 - 24, y1 + max_logo_h)
    img = fit_image(logo_for(sponsor["name"]), logo_box[2] - logo_box[0], logo_box[3] - logo_box[1])
    paste_centered(base, img, logo_box)

    label = label_for(sponsor)
    fnt = font(FONT_REG, 24)
    spacing = 8
    while tracking_width(draw, label, fnt, spacing) > (x2 - x1 - 32) and spacing > 1:
        spacing -= 1
    while tracking_width(draw, label, fnt, spacing) > (x2 - x1 - 32) and fnt.size > 18:
        fnt = font(FONT_REG, fnt.size - 1)
    draw_tracking_centered(draw, (x1 + x2) // 2, y2 - 42, label, fnt, GREY, spacing)


def draw_class_block(base: Image.Image, x: int, y: int, w: int, h: int, year: str) -> None:
    draw = ImageDraw.Draw(base)
    draw.rectangle((x, y, x + w, y + h), fill=BLUE_BLOCK)
    divider_x = x + w - 130
    draw.rectangle((divider_x, y - 3, divider_x + 5, y + h + 3), fill=GOLD)
    fnt_text = font(FONT_REG, 50)
    fnt_year = font(FONT_REG, 72)
    draw_tracking_centered(draw, x + (divider_x - x) // 2, y + 28, "CLASS OF", fnt_text, WHITE, 19)
    draw.text((divider_x + 20, y + 15), year, font=fnt_year, fill=(225, 181, 44))


def main() -> None:
    gold, bronze = read_sponsors()

    old = Image.open(ROOT / "SJBHS100Sponsors_from_drive.jpg").convert("RGBA")
    base = Image.new("RGBA", (CANVAS_W, CANVAS_H), WHITE + (255,))
    base.alpha_composite(old.crop((0, 0, CANVAS_W, 760)), (0, 0))

    draw_vertical_band(base, "GOLD", 61, 770, 1458, GOLD, 72)
    draw_vertical_band(base, "BRONZE", 61, 1510, 1935, BRONZE, 60)

    cells = [
        (165, 800, 760, 1070),
        (825, 800, 1420, 1070),
        (165, 1150, 760, 1420),
        (825, 1150, 1420, 1420),
    ]
    for sponsor, cell in zip(gold, cells):
        draw_sponsor_cell(base, sponsor, cell, 190)

    year = "97"
    for item in bronze:
        digits = "".join(ch for ch in item if ch.isdigit())
        if digits:
            year = digits[-2:]
            break
    draw_class_block(base, 380, 1635, 820, 120, year)

    draw = ImageDraw.Draw(base)
    small = font(FONT_REG, 28)
    draw_tracking_centered(draw, CANVAS_W // 2, 1795, "BATCH OF 1997", small, GREY, 10)

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    png_path = OUT_DIR / "SJBHS_sponsors_updated.png"
    jpg_path = OUT_DIR / "SJBHS_sponsors_updated.jpg"
    base.convert("RGB").save(jpg_path, quality=95, subsampling=0)
    base.save(png_path)
    print(jpg_path)
    print(png_path)


if __name__ == "__main__":
    main()
